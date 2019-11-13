#!/usr/bin/env python
#-*- coding: UTF-8 -*-

import wx
import os
import requests  # 导入requests
import re
import urllib.request
from openpyxl.reader.excel import load_workbook
import pymysql
import xlrd
import wx.grid
import math
import time
import datetime
import random
from googletrans import Translator
from w3lib.html import replace_entities
from urllib import error


#####################################Last Update: 2019-08-22#############################################################

class ButtonFrame(wx.Frame):




    def __init__(self):#####布局
        global  people

        #self.insert()
        #self.addTag()
        #self.trans_en()
        #self.transfer()
        #aself.correctStar()
        #self.addText()





        people = "Admin"

        mainTitle= '    ATS V2_822        Developed by TP-Link TSS Deutschland,         for: '+ people
        wx.Frame.__init__(self, None, -1, mainTitle,
                size=(700, 600))
        global panel

        panel = wx.Panel(self, -1)
        #####################一Note ################################
        self.text = wx.TextCtrl(panel,-1, value='Your Name', pos=(550,450), size=(100,30), style=0)
        self.text.SetBackgroundColour("#F0F0F0")

        self.Bind(wx.EVT_TEXT, self.onText, self.text)


        # global l1# 　　　静态文本　给用户提示信息的地方
        # l1 = wx.StaticText(panel, label="Note:       Please Load File", style=wx.ALIGN_CENTRE , size = (450,30),pos=(100, 180))
        # l1.SetForegroundColour('blue')
        # font = wx.Font(13,wx.DECORATIVE,wx.NORMAL,wx.NORMAL)
        # l1.SetFont(font)


        ################################operation buttons##########################################

        #################################admin buttons#############################################################
        #####################按钮1 update ################################
        self.button1 = wx.Button(panel, -1, "Update", pos=(550, 30), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.updateDB,self.button1)
        self.button1.SetDefault()

        #####################按钮2 save query to file################################
        self.button2 = wx.Button(panel, -1, "Query and Save", pos=(195, 30), size = (110,50) )
        self.Bind(wx.EVT_BUTTON, self.querySave,self.button2)
        self.button2.SetDefault()

        #####################按钮3 upload reply status################################
        self.button3 = wx.Button(panel, -1, "Set Reply Status", pos=(340, 30), size = (110,50) )
        self.Bind(wx.EVT_BUTTON, self.replyUpdate,self.button3)
        self.button3.SetDefault()


        ##################按钮4  add Product#############################
        self.button4 = wx.Button(panel, -1, "Add Product", pos=(550, 100), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.addProduct,self.button4)
        self.button4.SetDefault()

        ##################按钮5 add catalog #############################
        self.button5 = wx.Button(panel, -1, "Add Catalog", pos=(550, 170), size=(100, 50))
        self.Bind(wx.EVT_BUTTON, self.addCatalog, self.button5)
        self.button5.SetDefault()

        ##################按钮6 add Qestion#############################
        self.button6 = wx.Button(panel, -1, "Add Qestion", pos=(550, 240), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.addQestion,self.button6)
        self.button6.SetDefault()

        ##################按钮7 Show Question List#############################
        self.button7 = wx.Button(panel, -1, "Show List", pos=(550, 310), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.showList,self.button7)
        self.button7.SetDefault()

        ##################按钮8  database generate#############################
        self.button8 = wx.Button(panel, -1, "DB Generate", pos=(550, 380), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.updateDBFromFile,self.button8)
        self.button8.SetDefault()

        #####################按钮9 overview ################################
        self.button9 = wx.Button(panel, -1, "Overview", pos=(50, 30), size = (110,50) )
        self.Bind(wx.EVT_BUTTON, self.overview,self.button9)
        self.button9.SetDefault()

        #####################按钮10 refresh ################################
        self.button10 = wx.Button(panel, -1, "Refresh", pos=(550, 490), size = (100,50) )
        self.Bind(wx.EVT_BUTTON, self.refresh,self.button10)
        self.button10.SetDefault()



        ################GET DATABASE INFO#####################
        db = self.openDB()
        cur = db[0]
        conn = db[1]


        cur.execute('select max(date) from result ')
        dd = cur.fetchall()

        cur.close()
        conn.close()
        print("DB closed")


        # #########创建静态文本###############
        global statictext1,statictext2
        label2 ='____________________    Last Update:     %s    _________________________'%dd[0][0]
        label1 ='___________________________________________________________________________________'
        statictext2 = wx.StaticText(panel, label= label2,pos=(50, 90), size = (450,30))
        statictext1 = wx.StaticText(panel,
                                   label=label1,
                                   pos=(50, 180), size=(450, 30))


        ###########第1个下拉框 Catalog Name######
        wx.StaticText(panel, label='Catalog:', pos=(50, 120), size=(60, 30))
        list1 = ['All', 'Router', 'Modem','SMH','PLC','Repeater','AP','Mesh','Accessoire','Switch','Adapter','3G/4G','B.Router','Bundle','Neffos']
        self.ch1 = wx.ComboBox(panel, -1, value='All', choices=list1, style=wx.CB_SORT,pos=(50, 150), size = (70,30))
        self.Bind(wx.EVT_COMBOBOX, self.onCombo1, self.ch1)
        global queryValue1
        queryValue1 = 'All'

        ###########第2个下拉框 Star######

        wx.StaticText(panel, label='Star:', pos=(135, 120), size=(60, 30))
        list2 = ['All','1', '2','3', '4','5','Positiv','Negativ']
        self.ch2 = wx.ComboBox(panel, -1, value='All', choices=list2, style=wx.CB_SORT,pos=(135, 150), size = (65,30))
        self.Bind(wx.EVT_COMBOBOX, self.onCombo2, self.ch2)
        global queryValue2
        queryValue2 = 'All'


        ###########第3个下拉框 data from######
        global dateInput1
        wx.StaticText(panel, label='Date From:', pos=(210, 120), size=(80, 30))
        dateInput1 = wx.TextCtrl(panel, pos=(210, 150), size=(80, 25), style=0, name="")

        global dateInput2
        wx.StaticText(panel, label='Date To:', pos=(310, 120), size=(80, 30))
        dateInput2 = wx.TextCtrl(panel, pos=(310, 150), size=(80, 25), style=0, name="")

        wx.StaticText(panel, label='Replied ?', pos=(400, 120), size=(80, 30))
        list3 = ['All','No', 'Yes']
        self.ch3 = wx.ComboBox(panel, -1, value='All', choices=list3, style=wx.CB_SORT, pos=(400, 150), size=(50, 30))
        self.Bind(wx.EVT_COMBOBOX, self.onCombo3, self.ch3)
        global queryValue3
        queryValue3 = 'All'


    def onCombo1(self, event):
        global queryValue1
        queryValue1 ="All"
        queryValue1 = self.ch1.GetValue()
        #print(queryValue1)

    ###########第2个询关键字选择框返回值######
    def onCombo2(self, event):
        global queryValue2
        queryValue2 = "All"
        queryValue2 = self.ch2.GetValue()
        #print(self.phaseSelection)

    ###########第3个询关键字选择框返回值######
    def onCombo3(self, event):
        global queryValue3
        queryValue3 = "All"
        queryValue3= self.ch3.GetValue()
        #print(self.phaseSelection)

    def onText(self,event):
        global people
        people = self.text.GetValue()
        mainTitle = '    ATS V.2        Developed by TP-Link TSS Deutschland,         for: ' + people

        frame.SetTitle(mainTitle)






    ##########################Query and Save#############################################
    def querySave(self,event):
        path = self.open()
        wbSave = load_workbook(path)
        wsSave = wbSave.active


        db = self.openDB()
        cur = db[0]
        conn =db[1]

        #####输入日期 格式 2019-07-01
        dateBegin = dateInput1.GetValue()
        dateEnd = dateInput2.GetValue()
        if (dateEnd ==''):
            dateEnd = dateBegin


        catDict ={'All':'', 'Router':'and p_id >=1000 and p_id<=1999', 'Modem':'and p_id >=2000 and p_id<=2999','SMH':'and p_id >=3000 and p_id<=3999','PLC':'and p_id >=4000 and p_id<=4999','Repeater':'and p_id >=5000 and p_id<=5999','AP':'and p_id >=6000 and p_id<=6999','Mesh':'and p_id >=7000 and p_id<=7999','Accessoire':'and p_id >=8000 and p_id<=8999','Switch':'and p_id >=9000 and p_id<=9999','Adapter':'and p_id >=10000 and p_id<=10999','3G/4G':'and p_id >=11000 and p_id<=11999','B.Router':'and p_id >=12000 and p_id<=12999','Bundle':'and p_id >=13000 and p_id<=13999','Neffos':'and p_id >=14000 and p_id<=14999'}
        #print(queryValue1) #获取katagorie值
        value1 = catDict[queryValue1]#通过字典翻译成sql语句
        #print(value1)


        starDict ={'All':'','1':'and star=1', '2':'and star=2','3':'and star=3', '4':'and star=4','5':'and star=5','Positiv':'and star>3','Negativ':'and star<4'}#获取star值
        value2 = starDict[queryValue2]  # 通过字典翻译成sql语句
        print("value2",queryValue2,value2)


        replyDict ={'All':'','No':'and reply_status = 0', 'Yes':'and reply_status = 1'}#获取reply值
        value3 = replyDict[queryValue3]
        print(value3)# 通过字典翻译成sql语句




        sql = "select * from result where `date` >='%s' and `date` <= '%s' %s %s %s"%(dateBegin, dateEnd,value1,value2,value3)
        cur.execute(sql)
        outputQuery = cur.fetchall()
        #print(outputQuery)
        d1 = len(outputQuery)


        cur.execute('select max(date) from result ')
        d2 = cur.fetchall()

        cur.close()
        conn.close()
        print("DB closed")

        statictext1.SetLabel('___________________________%s  Records Selected__________________________________'%d1)
        statictext2.SetLabel('____________________    Last Update:     %s    _________________________'%d2[0][0])

        wsSave.cell(1, 1).value = 'Product ID'  # p_id
        wsSave.cell(1, 2).value = 'Product Name'  # p_name
        wsSave.cell(1, 3).value = 'Star' # star
        wsSave.cell(1, 4).value = 'Date'  # date
        wsSave.cell(1, 5).value =  'Reply'#reply status
        wsSave.cell(1, 6).value =  'Problem'#Problem
        wsSave.cell(1, 7).value = 'Username'  # user
        wsSave.cell(1, 8).value = 'Title'  # title
        wsSave.cell(1, 9).value = 'Review Link' # link
        wsSave.cell(1, 10).value = 'Review' # review
        wsSave.cell(1, 11).value = 'ID'  # review



        for i in range(d1):
            wsSave.cell(i + 2, 1).value = outputQuery[i][0]#p_id
            wsSave.cell(i + 2, 2).value = outputQuery[i][1]#p_name
            wsSave.cell(i + 2, 3).value = outputQuery[i][5]#star
            wsSave.cell(i + 2, 4).value = outputQuery[i][6]#date
            wsSave.cell(i + 2, 5).value = outputQuery[i][10]#Reply
            wsSave.cell(i + 2, 6).value = outputQuery[i][11]#Problem
            wsSave.cell(i + 2, 7).value = outputQuery[i][3]#user
            wsSave.cell(i + 2, 8).value = outputQuery[i][4]#title
            wsSave.cell(i + 2, 9).value = 'https://www.amazon.de/gp/customer-reviews/'+outputQuery[i][7]#link
            wsSave.cell(i + 2, 10).value = outputQuery[i][8]#review
            wsSave.cell(i + 2, 11).value = outputQuery[i][9]  # ID


        wbSave.save(path)





    ##########################添加产品信息#############################################
    def addProduct(self, event):
        path = self.open()

        try:

            db = self.openDB()
            cur = db[0]
            conn = db[1]



            wdata = xlrd.open_workbook(path)  ###这个库的表格单元是从[0][0]开始的
            table = wdata.sheets()[0]
            nrows = table.nrows

            for i in range(1,nrows):
                r1 = table.row_values(i)  # 从第2行开始读取，第一行是标题
                cur = conn.cursor()
                cur.execute('insert ignore into product values (%s,%s,%s)',
                            (r1[0], r1[1], r1[2]))
                conn.commit()
                cur.close()
            conn.close()
            print("DB closed")

        except Exception as e:
            print(e)

        print('success')






    ##############################添加目录产品#########################################
    def addCatalog(self, event):
        path = self.open()

        try:

            db = self.openDB()
            cur = db[0]
            conn = db[1]

            wdata = xlrd.open_workbook(path)  ###这个库的表格单元是从[0][0]开始的
            table = wdata.sheets()[0]
            nrows = table.nrows

            for i in range(1,nrows):
                r1 = table.row_values(i)  # 从第2行开始读取，第一行是标题
                cur = conn.cursor()
                cur.execute('insert ignore into category values (%s,%s)',
                            (r1[0], r1[1]))
                conn.commit()
                cur.close()
            conn.close()
            print("DB closed")

        except Exception as e:
            print(e)



    ###############################添加问题分类########################################
    def addQestion(self, event):
        path = self.open()

        try:

            db = self.openDB()
            cur = db[0]
            conn = db[1]

            wdata = xlrd.open_workbook(path)  ###这个库的表格单元是从[0][0]开始的
            table = wdata.sheets()[0]
            nrows = table.nrows

            for i in range(1, nrows):
                r1 = table.row_values(i)  # 从第2行开始读取，第一行是标题
                cur = conn.cursor()
                cur.execute('select count(1) from question')
                data=cur.fetchall()
                cur.execute('insert ignore into question values (%s,%s,%s,%s)',
                            (r1[0], r1[1], r1[2],data[0][0]+1))
                conn.commit()
                cur.close()
            conn.close()
            print("DB closed")

        except Exception as e:
            print(e)



    ##################################客户端抓取亚马逊信息 并 更新到数据库####################################
    def updateDBFromFile(self,event):
        start = time.clock()

        ########引入数据源表格 ，excel表格有固定格式， 第一行为标题（p_id,p_name,asin）,数据从第二行开始
        path = self.open()
        wb1 = load_workbook(path)
        ws1 = wb1.active

        # ###############引入数据源存储表格############
        # currentPath = r"C:\Users\dream\OneDrive\Desktop\amazon"
        #
        # resultFilePass = currentPath + r'\result.xlsx'  # excel operation save to result file
        # wb2 = load_workbook(resultFilePass)
        # ws2 = wb2.active

        nrProduct = ws1.max_row
        print(nrProduct)



        # 添加头部，伪装浏览器



        for j in range(2, nrProduct+1):  ###source表格数据是从第二行开始，到max+1行 结束

            # 获取数据库连接
            db = self.openDB()
            cur = db[0]
            conn = db[1]

            url = "https://www.amazon.de/product-reviews/"+ ws1.cell(j, 3).value+"/ref=cm_cr_arp_d_viewopt_fmt?formatType=current_format&pageNumber=1&sortBy=recent" ##asin是第3列数据

            print(url)
            # headers在这里不必须，嗯，还是加上吧...

            browser = str(random.randint(1, 4))
            version = str(random.randint(1, 10))
            word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (browser, version)



            headers = {
                'User-Agent': word}
            # Request类的实例，构造时需要传入Url,Data，headers等等的内容
            resquest = urllib.request.Request(url=url, headers=headers)
            response = urllib.request.urlopen(resquest).read()
            html = response.decode('utf-8')

            #print(html)

            count = re.findall('class="a-size-base\">.*?von (.*?) Rezensionen werden angezeigt', html)  # get count 亚马逊评论总数量
            #print(html)
            if (len(count)==0):#没有任何评论，跳过此次循环，记录count =0
                cur.execute('insert ignore  into count values (%s,%s)',
                            (ws1.cell(j, 1).value, 0))
                conn.commit()
                continue
            print(count," Rezensionen werden angezeigt")
            count = count[0]
            #ss= re.search("a-size-base\">.*?von ",count)
            #count =count.replace(ss.group(),"")
            count = count.replace(".", "")#超过1000的评论数 会产生 句号分隔
            #count = count.replace("von ", "")
            #count = count.replace(" Rezensionen werden angezeigt", "")
            print(count)

            count = int(count)#获取到评论数量

            cur.execute('insert ignore  into count values (%s,%s)',
                        (ws1.cell(j,1).value,count))

            conn.commit()

            list = [ [0 for col in range(9)] for row in range(count)]

            #print(len(list))


            for r in range(1,math.ceil(count/10)+1):  ## r:某一产品的页数

                page = str(r)
                url = "https://www.amazon.de/product-reviews/"+ws1.cell(j, 3).value+"/ref=cm_cr_arp_d_viewopt_fmt?formatType=current_format&pageNumber=%s&sortBy=recent"%(page)
                print(url)


                browser = str(random.randint(1, 4))
                version = str(random.randint(1, 10))
                word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (browser, version)

                headers = {
                    'User-Agent':word}
                # Request类的实例，构造时需要传入Url,Data，headers等等的内容
                resquest = urllib.request.Request(url=url, headers=headers)
                response = urllib.request.urlopen(resquest).read()
                html = response.decode('utf-8')

                tt = random.uniform(4, 7)
                time.sleep(tt)

                ####search

                aa1 = ws1.cell(j, 3).value + "\"><span class=\"\">.*?</span>"  # 获取asin+html的内容 -> title
                aa2 = ws1.cell(j, 3).value  # asin

                # <span class="a-profile-name">Paul Mierau</span>
                username = re.findall('a-profile-name.*?</span>', html)  # get username

                ## ASIN=B00A0VCJPI"><span class="">Device delivered with defect, no money back.</span>  #
                title = re.findall(aa1, html)  # get title

                ##<a class="a-link-normal" title="2,0 von 5 Sternen" href=
                star = re.findall('<a class=\"a-link-normal\" title=\".*?\" href=\"/gp', html)  ## get star

                # <span data-hook="review-date" class="a-size-base a-color-secondary review-date">13. Juni 2019</span>
                date = re.findall('a-size-base a-color-secondary review-date\">.*?</span>', html)  # get date

                # <span data-hook="review-body" class="a-size-base review-text review-text-content"><span class="">I ordered thisshes....sending the old back.<br /><br />router.</span>
                review = re.findall('a-size-base review-text review-text-content\"><span class=\"\">.*?</span>',
                                    html)  # get review

                # div id="customer_review-R29NTLZWDQTZLR" class
                review_id = re.findall('div id=\"customer_review-.*?\"', html)

                positiveReview = re.findall('Am höchsten bewertete positive Rezension', html)
                negativeReview = re.findall('Am höchsten bewertete kritische Rezension', html)
                print(positiveReview, len(positiveReview))




                nrOfRec = len(star)  # 根据评论数量，需要写入的行数
                print("nrOfRec is:",nrOfRec)


                ###如果被amazon屏蔽了，那同一个页面再运行一次
                t = 1
                while (nrOfRec==0 and r< (math.ceil(count/10)) and t<5):
                    browser = str(random.randint(1, 4))
                    version = str(random.randint(1, 10))
                    word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (
                    browser, version)

                    headers = {
                        'User-Agent': word}

                    time.sleep(10*t)

                    # Request类的实例，构造时需要传入Url,Data，headers等等的内容
                    resquest = urllib.request.Request(url=url, headers=headers)
                    response = urllib.request.urlopen(resquest).read()
                    html = response.decode('utf-8')



                    ####search

                    aa1 = ws1.cell(j, 3).value + "\"><span class=\"\">.*?</span>"  # 获取asin+html的内容 -> title
                    aa2 = ws1.cell(j, 3).value  # asin


                    username = re.findall('a-profile-name.*?</span>', html)  # get username


                    title = re.findall(aa1, html)  # get title


                    star = re.findall('<a class=\"a-link-normal\" title=\".*?\" href=\"/gp', html)  ## get star


                    date = re.findall('a-size-base a-color-secondary review-date\">.*?</span>', html)  # get date


                    review = re.findall('a-size-base review-text review-text-content\"><span class=\"\">.*?</span>',
                                        html)  # get review


                    review_id = re.findall('div id=\"customer_review-.*?\"', html)


                    positiveReview = re.findall('Am höchsten bewertete positive Rezension',html)
                    negativeReview = re.findall('Am höchsten bewertete kritische Rezension',html)
                    print(positiveReview,len(positiveReview))

                    nrOfRec = len(star)  # 根据评论数量，需要写入的行数
                    print("nrOfRec is:", nrOfRec)

                if (nrOfRec == 0):####尝试多次以后 仍然没有抓取到数据 break
                    print("ASIN:  "+ws1.cell(j, 3)+" failed "+" Page:  "+page)
                    break


                # print(user)
                # print(startRow,nrOfRec,startRow+nrOfRec)

                #####写入第j行的第i列
                for i in range(nrOfRec):##i:每页有10条评论

                    # write title
                    title[i] = title[i].replace(aa2, "")
                    title[i] = title[i].replace("\"><span class=\"\">", "")
                    title[i] = title[i].replace("</span>", "")
                    #ws2.cell(i, 6).value = title[i - startRow]
                    title[i] = replace_entities(title[i])


                    # write stars
                    star[i] = star[i].replace("<a class=\"a-link-normal\" title=\"", "")
                    star[i] = star[i].replace(" von 5 Sternen\" href=", "")
                    #ws2.cell(i, 7).value = star[i - startRow]

                    # write review_id
                    review_id[i] = review_id[i].replace("div id=\"customer_review-", "")
                    review_id[i] = review_id[i].replace("\"", "")
                    #ws2.cell(i, 9).value = review_id[i - startRow]

                    # write review
                    review[i] = review[i].replace(
                        "a-size-base review-text review-text-content\"><span class=\"\">", "")
                    review[i] = review[i].replace("</span>", "")
                    review[i] = review[i].replace("<br />", "")
                    #ws2.cell(i, 10).value = review[i - startRow]
                    review[i] = replace_entities(review[i])


                    b = i+2
                    if (len(positiveReview)==0):# 引入b，因为有的页面没有最佳和最差，导致排序往前
                        b = i

                    # 写入user
                    username[b] = username[b].replace("a-profile-name\">","")  # 因为用户会多出两个，所以要除去多出的两个，从第3个开始一直到第12个，提取html信息
                    username[b] = username[b].replace("</span>", "")  # 去除标签
                    #ws2.cell(i, 5).value = username[i - startRow + 2]  # save to result
                    username[b] = replace_entities(username[b])

                    # write date
                    date[b] = date[b].replace("a-size-base a-color-secondary review-date\">",
                                                                            "")
                    date[b] = date[b].replace("</span>", "")
                    date[b] = date[b].replace('.', '')

                    ## 对日期进行格式转换####
                    date[b] = date[b].split(' ')

                    dayDict = {'1': '01', '2': '02', '3': '03', '4': '04', '5': '05', '6': '06', '7': '07', '8': '08',
                               '9': '09'}
                    dateDict = {'Januar': '01', 'Februar': '02', 'März': '03', 'April': '04', 'Mai': '05', 'Juni': '06',
                                'Juli': '07', 'August': '08', 'September': '09', 'Oktober': '10', 'November': '11',
                                'Dezember': '12'}
                    if date[b][0] in dayDict:
                        date[b][0] = dayDict[date[b][0]]
                    date[b][1] = dateDict[date[b][1]]
                    word = date[b][2] + '-' + date[b][1] + '-' + date[b][0]
                    ###########日期格式转换结束###############



                    list[(r - 1) * 10  + i] = [ws1.cell(j, 1).value, ws1.cell(j, 2).value,
                                                             ws1.cell(j, 3).value,
                                                             username[b], title[i],
                                                             star[i], word, review_id[i],review[i]]
                    print((r - 1) * 10 + i,"page:",r, "P_nr:",ws1.cell(j,1).value )
                    #print(list)
                    #print("pause")
                    #print(len(list))




            ######################将一种产品的信息一次存入数据库   开始########################

            for s in range(len(list)):

                if(list[s][0]==0): ##防止list中的这条数据为空，空记录造成原因是amazon的评论数量不准确，导致实际评论数量少于显示的数量
                    continue

                cur.execute('select count(1) from  result;')
                data = cur.fetchall()
                print(data[0][0])
                # cur.execute('alter table tplink.result AUTO_INCREMENT= %s', (data[0][0] + 1))
                cur.execute('insert ignore  into result values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0)',
                            (list[s][0], list[s][1], list[s][2], list[s][3], list[s][4], list[s][5], list[s][6], list[s][7], list[s][8],data[0][0]+1))

                conn.commit()

            cur.close()
            conn.close()
            print("DB closed")


            ##########################将一种产品的信息一次存入数据库  结束#############################


        print("success!")
        end = time.clock()
        print(str(end - start))

    ################################打开文件##################################
    def open(self):

        global wordPath0
        dlg = wx.FileDialog(self, message=u"Choose  Source-File Please",
                            defaultDir=os.getcwd(),
                            defaultFile="")

        if dlg.ShowModal() == wx.ID_OK:

            wordPaths0 = dlg.GetPaths()  # 返回一个list，如[u'E:\\test_python\\Demo\\ColourDialog.py', u'E:\\test_python\\Demo\\DirDialog.py']
            #print wordPaths
            for wordPath0 in wordPaths0:
                print(wordPath0)  #  打开的文件地址

        return wordPath0

        # global l2# 显示打开的文件地址
        # l2 = wx.StaticText(panel, label="Source File:  "+ wordPath0, style=wx.ALIGN_CENTRE , size = (450,30),pos=(120, 100))
        # l2.SetForegroundColour('grey')

    ##################################更新数据库###############################
    def updateDB(self,event):

        db = self.openDB()
        cur = db[0]
        conn = db[1]

        cur.execute('select count(p_id) from  product')
        dataNr = cur.fetchall()#产品总数
        #print(data[0][0])
        cur.close()
        conn.close()
        print("DB closed")

        for nr in range (380,dataNr[0][0]):
            db = self.openDB()
            cur = db[0]
            conn = db[1]

            cur.execute('select * from  product limit %s,1',nr)
            dataP=cur.fetchall()
            print(dataP[0][0])#p_id
            print(dataP[0][1])#p_name
            print(dataP[0][2])#asin

            cur.execute('select amount from  count where p_id = %s', dataP[0][0])
            countOld = cur.fetchall()
            countOld = countOld[0][0]
            print("countOld = ", countOld)
            print(dataP[0][0])



            url = "https://www.amazon.de/product-reviews/" + dataP[0][2] + "/ref=cm_cr_arp_d_viewopt_fmt?formatType=current_format&pageNumber=1&sortBy=recent"  ##asin是第3列数据
            print(url)


            browser = str(random.randint(1, 4))
            version = str(random.randint(1, 10))
            word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (
            browser, version)

            headers = {
                'User-Agent': word}
            # Request类的实例，构造时需要传入Url,Data，headers等等的内容
            resquest = urllib.request.Request(url=url, headers=headers)
            try:
                response = urllib.request.urlopen(resquest).read()
            except error.URLError as e:

                print("HTTPError")
                continue



            html = response.decode('utf-8')

            # print(html)


            countNew = re.findall('class="a-size-base\">.*?von (.*?) Rezensionen werden angezeigt',
                               html)  # get count 亚马逊评论总数量
            # print(html)
            if (len(countNew) == 0):  # 没有任何评论，跳过此次循环，记录count =0
                continue
            print(countNew, " Rezensionen werden angezeigt")
            countNew = countNew[0]
            countNew = countNew.replace(".", "")  # 超过1000的评论数 会产生 句号分隔
            print("countNew = ",countNew)
            countNew = int(countNew)

            countNeed = min( (countNew - countOld),countNew)
            print("countNeed = ",countNeed)

            list = [[0 for col in range(9)] for row in range(countNeed)]
            restPosition = countNeed

            for r in range(1, math.ceil(countNeed / 10) + 1):  ## r:某一产品的页数,从1开始

                page = str(r)
                url = "https://www.amazon.de/product-reviews/" + dataP[0][2]+ "/ref=cm_cr_arp_d_viewopt_fmt?formatType=current_format&pageNumber=%s&sortBy=recent" % (page)
                print(url)

                browser = str(random.randint(1, 4))
                version = str(random.randint(1, 10))
                word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (
                browser, version)

                headers = {
                    'User-Agent': word}
                # Request类的实例，构造时需要传入Url,Data，headers等等的内容
                resquest = urllib.request.Request(url=url, headers=headers)
                try:
                    response = urllib.request.urlopen(resquest).read()
                except error.URLError as e:

                    print("HTTPError")
                    continue

                html = response.decode('utf-8')

                tt = random.uniform(4, 5)
                time.sleep(tt)

                ####search

                aa1 = dataP[0][2]+ "\"><span class=\"\">.*?</span>"  # 获取asin+html的内容 -> title
                aa2 = dataP[0][2] # asin

                # <span class="a-profile-name">Paul Mierau</span>
                username = re.findall('a-profile-name.*?</span>', html)  # get username

                ## ASIN=B00A0VCJPI"><span class="">Device delivered with defect, no money back.</span>  #
                title = re.findall(aa1, html)  # get title

                ##<a class="a-link-normal" title="2,0 von 5 Sternen" href=
                star = re.findall('<a class=\"a-link-normal\" title=\".*?\" href=\"/gp', html)  ## get star

                # <span data-hook="review-date" class="a-size-base a-color-secondary review-date">13. Juni 2019</span>
                date = re.findall('a-size-base a-color-secondary review-date\">.*?</span>', html)  # get date

                # <span data-hook="review-body" class="a-size-base review-text review-text-content"><span class="">I ordered thisshes....sending the old back.<br /><br />router.</span>
                review = re.findall('a-size-base review-text review-text-content\"><span class=\"\">.*?</span>',
                                    html)  # get review

                # div id="customer_review-R29NTLZWDQTZLR" class
                review_id = re.findall('div id=\"customer_review-.*?\"', html)

                positiveReview = re.findall('Am höchsten bewertete positive Rezension', html)
                negativeReview = re.findall('Am höchsten bewertete kritische Rezension', html)
                print(positiveReview, len(positiveReview))

                nrOfRec = len(star)  # 根据评论数量，需要写入的行数
                print("nrOfRec is:", nrOfRec)

                ###如果被amazon屏蔽了，那同一个页面再运行一次
                t = 1
                while (nrOfRec == 0 and r < (math.ceil(countNeed / 10)) and t < 5):
                    browser = str(random.randint(1, 4))
                    version = str(random.randint(1, 10))
                    word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (
                        browser, version)

                    headers = {
                        'User-Agent': word}

                    time.sleep(10 * t)

                    # Request类的实例，构造时需要传入Url,Data，headers等等的内容
                    resquest = urllib.request.Request(url=url, headers=headers)
                    try:
                        response = urllib.request.urlopen(resquest).read()
                    except error.URLError as e:

                        print("HTTPError")

                    html = response.decode('utf-8')

                    ####search

                    aa1 = dataP[0][2]+ "\"><span class=\"\">.*?</span>"  # 获取asin+html的内容 -> title
                    aa2 = dataP[0][2]  # asin

                    username = re.findall('a-profile-name.*?</span>', html)  # get username

                    title = re.findall(aa1, html)  # get title

                    star = re.findall('<a class=\"a-link-normal\" title=\".*?\" href=\"/gp', html)  ## get star

                    date = re.findall('a-size-base a-color-secondary review-date\">.*?</span>', html)  # get date

                    review = re.findall('a-size-base review-text review-text-content\"><span class=\"\">.*?</span>',
                                        html)  # get review

                    review_id = re.findall('div id=\"customer_review-.*?\"', html)

                    positiveReview = re.findall('Am höchsten bewertete positive Rezension', html)
                    negativeReview = re.findall('Am höchsten bewertete kritische Rezension', html)
                    print(positiveReview, len(positiveReview))

                    nrOfRec = len(star)  # 根据评论数量，需要写入的行数
                    print("nrOfRec is:", nrOfRec)

                if (nrOfRec == 0):  ####尝试多次以后 仍然没有抓取到数据 break
                    print("ASIN:  " + dataP[0][2]+ " failed " + " Page:  " + page)
                    break

                # print(user)
                # print(startRow,nrOfRec,startRow+nrOfRec)
                nrOfRec = min(restPosition,nrOfRec,10)
                print("nrOfRec is:", nrOfRec)

                #####写入第j行的第i列
                for i in range(nrOfRec):  ##i:每页有10条评论

                    restPosition = restPosition -1
                    print("i = ",i)


                    title[i] = title[i].replace(aa2, "")
                    title[i] = title[i].replace("\"><span class=\"\">", "")
                    title[i] = title[i].replace("</span>", "")
                    # ws2.cell(i, 6).value = title[i - startRow]
                    title[i] = replace_entities(title[i])

                    # write stars
                    star[i] = star[i].replace("<a class=\"a-link-normal\" title=\"", "")
                    star[i] = star[i].replace(" von 5 Sternen\" href=", "")
                    # ws2.cell(i, 7).value = star[i - startRow]

                    # write review_id
                    review_id[i] = review_id[i].replace("div id=\"customer_review-", "")
                    review_id[i] = review_id[i].replace("\"", "")
                    # ws2.cell(i, 9).value = review_id[i - startRow]

                    # write review
                    review[i] = review[i].replace(
                        "a-size-base review-text review-text-content\"><span class=\"\">", "")
                    review[i] = review[i].replace("</span>", "")
                    review[i] = review[i].replace("<br />", "")
                    # ws2.cell(i, 10).value = review[i - startRow]
                    review[i] = replace_entities(review[i])

                    b = i + 2

                    if (len(positiveReview) == 0):  # 引入b，因为有的页面没有最佳和最差，导致排序往前
                        b = i

                    # 写入user
                    username[b] = username[b].replace("a-profile-name\">",
                                                      "")  # 因为用户会多出两个，所以要除去多出的两个，从第3个开始一直到第12个，提取html信息
                    username[b] = username[b].replace("</span>", "")  # 去除标签
                    username[b] = replace_entities(username[b])

                    # write date
                    date[b] = date[b].replace("a-size-base a-color-secondary review-date\">",
                                              "")
                    date[b] = date[b].replace("</span>", "")
                    date[b] = date[b].replace('.', '')

                    ## 对日期进行格式转换####
                    date[b] = date[b].split(' ')

                    dayDict = {'1': '01', '2': '02', '3': '03', '4': '04', '5': '05', '6': '06', '7': '07', '8': '08',
                               '9': '09'}
                    dateDict = {'Januar': '01', 'Februar': '02', 'März': '03', 'April': '04', 'Mai': '05', 'Juni': '06',
                                'Juli': '07', 'August': '08', 'September': '09', 'Oktober': '10', 'November': '11',
                                'Dezember': '12'}
                    if date[b][0] in dayDict:
                        date[b][0] = dayDict[date[b][0]]
                    date[b][1] = dateDict[date[b][1]]
                    word = date[b][2] + '-' + date[b][1] + '-' + date[b][0]
                    ###########日期格式转换结束###############


                    list[(r - 1) * 10 + i] = [dataP[0][0], dataP[0][1],
                                              dataP[0][2],
                                              username[b], title[i],
                                              star[i], word, review_id[i], review[i]]
                    print((r - 1) * 10 + i, "page:", r, "P_nr:", dataP[0][0],"  RestPosition =  ",restPosition)
                    # print(list)
                    # print("pause")
                    # print(len(list))


            ######################将一种产品的信息一次存入数据库   开始########################

            for s in range(len(list)):

                if (list[s][0] == 0):  ##防止list中的这条数据为空，空记录造成原因是amazon的评论数量不准确，导致实际评论数量少于显示的数量
                    continue

                cur.execute('select count(1) from  result;')
                data = cur.fetchall()
                print(data[0][0])
                # cur.execute('alter table tplink.result AUTO_INCREMENT= %s', (data[0][0] + 1))
                cur.execute('insert ignore  into result values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,0,0,0)',
                            (list[s][0], list[s][1], list[s][2], list[s][3], list[s][4], list[s][5], list[s][6],
                             list[s][7], list[s][8], data[0][0] + 1))

                cur.execute('UPDATE `amazon`.`count` SET `amount` = %s WHERE (`p_id` = %s)',(countNew,dataP[0][0]))#更新回复数量

                conn.commit()

            cur.close()
            conn.close()
            print("DB closed")

            ##########################将一种产品的信息一次存入数据库  结束#############################

        print("success!")





    ##############################reply update################################################
    def replyUpdate(self,event):
        global people,theTime
        ISOTIMEFORMAT = '%Y-%m-%d'
        theTime = datetime.datetime.now().strftime(ISOTIMEFORMAT)

        db = self.openDB()
        cur = db[0]
        conn = db[1]

        path=self.open()

        wbUpdate = load_workbook(path)
        wsUpdate = wbUpdate.active

        rows = wsUpdate.max_row
        print(rows)

        for i in range(2,rows+1):


            if (wsUpdate.cell(i,5).value==0):
                continue

            cur.execute( 'UPDATE result SET qes_id = %s ,reply_status = %s , reply_date = %s , owner = %s where id = %s',(wsUpdate.cell(i,6).value, wsUpdate.cell(i,5).value,theTime,people, wsUpdate.cell(i,11).value))




        conn.commit()
        cur.close()
        conn.close()
        print("DB closed")

    def refresh(self,event):
        print()
        global  ctrl2
        if (ctrl2==1):
            grid2.Destroy()
            ctrl2=0
        global  ctrl
        if (ctrl==1):
            grid.Destroy()
            ctrl=0




    ##########################overview: grid#####################################
    global ctrl,ctrl2
    ctrl=0
    ctrl2=0

    def overview(self,event):

        global  ctrl,ctrl2,grid
        if (ctrl2==1):
            grid2.Destroy()
            ctrl2=0

        if (ctrl==1):
            grid.Destroy()
            ctrl=0

        db = self.openDB()
        cur = db[0]
        conn = db[1]


        dateBegin = dateInput1.GetValue()
        dateEnd = dateInput2.GetValue()
        if (dateEnd == ''):
            dateEnd = dateBegin

        overviewList1 = ['and p_id >=1000 and p_id<=1999', 'and p_id >=2000 and p_id<=2999',
                         'and p_id >=3000 and p_id<=3999', 'and p_id >=4000 and p_id<=4999',
                         'and p_id >=5000 and p_id<=5999', 'and p_id >=6000 and p_id<=6999',
                         'and p_id >=7000 and p_id<=7999', 'and p_id >=8000 and p_id<=8999',
                         'and p_id >=9000 and p_id<=9999', 'and p_id >=10000 and p_id<=10999',
                         'and p_id >=11000 and p_id<=11999', 'and p_id >=12000 and p_id<=12999',
                         'and p_id >=13000 and p_id<=13999', 'and p_id >=14000 and p_id<=14999']

        overviewList2 = ['and star=1', 'and star=2', 'and star=3', 'and star>3', 'and star<4',
                         'and reply_status = 1', '']



        resultOV = [[0 for col in range(7)] for row in range(14)]
        for i in range(14):
            for j in range(7):


                sql = "select count(p_id) from `result` where `date` >= '%s' and `date` <= '%s' %s %s" % (
                dateBegin, dateEnd, overviewList1[i], overviewList2[j])

                cur.execute(sql)
                data = cur.fetchall()
                resultOV[i][j] = str(data[0][0])

        cur.close()
        conn.close()
        print('DB closed')



        grid = wx.grid.Grid(panel, pos=(50, 220), size=(450, 350))  # 创建grid视图

        grid.SetMargins(-200, -500)
        grid.CreateGrid(14, 8)  # 建表

        grid.SetDefaultCellBackgroundColour("#F0F0F0")

        grid.SetColLabelValue(0, "Cata.")  # 初始化列标题
        grid.SetColSize(col=0, width=60)  # 初始化大小

        grid.SetColLabelValue(1, "1")  # 初始化列标题
        grid.SetColSize(col=1, width=40)  # 初始化大小

        grid.SetColLabelValue(2, "2")  # 初始化列标题
        grid.SetColSize(col=2, width=40)  # 初始化大小

        grid.SetColLabelValue(3, "3")  # 初始化列标题
        grid.SetColSize(col=3, width=40)  # 初始化大小

        grid.SetColLabelValue(4, "Pos.")  # 初始化列标题
        grid.SetColSize(col=4, width=40)  # 初始化大小

        grid.SetColLabelValue(5, "Neg.")  # 初始化列标题
        grid.SetColSize(col=5, width=40)  # 初始化大小

        grid.SetColLabelValue(6, "Rpl.")  # 初始化列标题
        grid.SetColSize(col=6, width=40)  # 初始化大小
        grid.SetColLabelValue(7, "Σ")  # 初始化列标题
        grid.SetColSize(col=7, width=40)  # 初始化大小

        cataList = ['Router','Modem','SMH','PLC','Re.','AP','Mesh','Access.','Switch','Adapter','3G/4G','B.Router','Bundle','Neffos']
        for i in range(14):

            grid.SetCellValue(i, 0, cataList[i])


        for i in range(0,14):
            for j in range(1,8):
                grid.SetCellValue(i, j, resultOV[i][j-1])

        ctrl= 1






    ###################show list:grid2##############################


    def showList(self,event):
        global ctrl,ctrl2,grid2


        if (ctrl ==1):
            grid.Destroy()
            ctrl =0

        if (ctrl2 == 1):
            grid2.Destroy()
            ctrl2 = 0
        print(queryValue1)

        dictQueryValue1 = {'All':0, 'Router':1, 'Modem':2,'SMH':3,'PLC':4,'Repeater':5,'AP':6,'Mesh':7,'Accessoire':0,'Switch':9,'Adapter':10,'3G/4G':11,'B.Router':12,'Bundle':0,'Neffos':0}

        db = self.openDB()
        cur = db[0]
        conn = db[1]
        aa = dictQueryValue1[queryValue1]
        print(aa)
        sql = 'select qes_id,qes from question where cat_id = %s ORDER BY qes_id'%aa
        cur.execute(sql)
        data = cur.fetchall()
        print(len(data))
        print(data)
        cur.close()
        conn.close()
        print('DB closed')

        if (aa !=0):



            grid2 = wx.grid.Grid(panel, pos=(50, 220), size=(450, 350))  # 创建grid视图

            grid2.SetMargins(-200, -500)
            grid2.CreateGrid(len(data), 2)  # 建表

            grid2.SetDefaultCellBackgroundColour("#F0F0F0")

            grid2.SetColLabelValue(0, "Problem ID.")  # 初始化列标题
            grid2.SetColSize(col=0, width=80)  # 初始化大小

            grid2.SetColLabelValue(1, "Problem")  # 初始化列标题
            grid2.SetColSize(col=1, width=200)  # 初始化大小



            for i in range(0, len(data)):
                for j in range(2):
                    grid2.SetCellValue(i, j, str(data[i][j]))

            #grid2.Shown

            ctrl2 = 1

        print(people)



    ###################open database##############
    def openDB(self):
        global cur,conn
        conn = pymysql.connect(host='172.31.1.1', user='root', port=3306,password='L1i2n3k4!')
        # 打印数据库连接对象
        #print('数据库连接对象为：{}'.format(conn))
        # 获取游标
        cur = conn.cursor()
        sql = "use amazon "
        cur.execute(sql)
        print("DB Connected and "+ sql)
        return cur, conn

    ###################Translate to EN##############
    ##Trans DE to EN
    def trans_en(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]

        sql = 'select id,review_de from translation where id >= 37067 and id<= 37714 '
        cur.execute(sql)
        data = cur.fetchall()

        translator = Translator()
        for i in range (10000):
            print(data[i][0])
            print(data[i][1])
            s = data[i][1]

            s = replace_entities(s)
            s = re.sub(r'[^\w\s+\.\!\/_,$%^*(+\"\']+|[+——！，。？、~@#￥%……&*（）：]', "", s)
            t = translator.translate(s, 'EN')

            print(t.text)
            cur.execute('UPDATE translation SET review_en = %s  WHERE id = %s;', (t.text,data[i][0]))

            tt = random.uniform(3, 7)
            time.sleep(tt)

            conn.commit()



        cur.close()
        conn.close()


    ###################Transfer : move the reviews of EN from result_table to translation_table##############
    def transfer(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]

        cur.execute('select count(1) from result')
        num = cur.fetchall()[0][0]
        print(num)

        sql = 'select id,review from result where result.id<=37714 and result.id>=37067'
        cur.execute(sql)
        data = cur.fetchall()


        for i in range(1509):
            print(data[i][0])
            print(data[i][1])
            s = data[i][1]
            s = replace_entities(s)


            cur.execute('insert ignore into translation values (%s,%s,0,0)', (data[i][0], s))

            conn.commit()
        cur.close()
        conn.close()

    ###################Correct: encode html special symbol to utf-8##############
    def correct(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]

        cur.execute('select count(1) from result')
        num = cur.fetchall()[0][0]
        print(num)

        sql = 'select id,username,title,review from result'
        cur.execute(sql)
        data = cur.fetchall()

        for i in range(num):
            print(data[i][0])

            s0 = data[i][0]
            s1 = replace_entities(data[i][1])
            s2 = replace_entities(data[i][2])
            s3 = replace_entities(data[i][3])

            cur.execute('UPDATE result SET username = %s , title = %s,review = %s  WHERE id = %s;', (s1,s2,s3, s0))

            conn.commit()
        cur.close()
        conn.close()

    def correctStar(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]

        for i in range(37419,37715):

            cur.execute('select review_id from result where id=%s;', i)
            data = cur.fetchall()
            print(data)

            url = "https://www.amazon.de/gp/customer-reviews/" + data[0][0]
            print(url)

            browser = str(random.randint(1, 4))
            version = str(random.randint(1, 10))
            word = 'Mozilla/%s.0 (Windows NT %s.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.103 Safari/537.36' % (
                browser, version)

            headers = {
                'User-Agent': word}
            # Request类的实例，构造时需要传入Url,Data，headers等等的内容
            resquest = urllib.request.Request(url=url, headers=headers)
            try:
                response = urllib.request.urlopen(resquest).read()
            except error.URLError as e:

                print("HTTPError")
                continue

            html = response.decode('utf-8')

            # print(html)
            print('id = ',i)
            star = re.findall('<a class=\"a-link-normal\" title=\".*?\" href=\"/gp', html)
            print('star = ', star)

            star = star[0].replace("<a class=\"a-link-normal\" title=\"", "")
            star = star[0].replace(" von 5 Sternen\" href=\"/gp", "")
            print('star = ', star)
            print(type(star))
            cur.execute('UPDATE result SET star = %s where id = %s',(star,i))
            conn.commit()
            tt = random.uniform(1, 2)
            time.sleep(tt)
        cur.close()
        conn.close()

    def insert(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]

        cur.execute('select count(1) from repeater')
        num = cur.fetchall()[0][0]
        print(num)
        cur.execute('select id from repeater')
        id = cur.fetchall()
        print(id)
        print(id[0][0])
        for i in range(num):
            print(i)
            cur.execute('select review_en from translation where id = %s',id[i][0])
            en = cur.fetchall()[0][0]
            print(en)

            cur.execute('UPDATE repeater SET review_en = %s where id = %s', (en, id[i][0]))
            conn.commit()
        cur.close()
        conn.close()

    def addTag(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]
        #cur.execute('UPDATE repeater SET tag = %s where star <= 5 ',('ohters'))

        #cur.execute('UPDATE repeater SET tag = %s where star <= 3 and review_en like %s or %s', ('no internet','%no internet%','%no internet%'))
        cur.execute('select count(1) from repeater   ')
        count = cur.fetchall()[0][0]
        print(count)
        conn.commit()
        cur.close()
        conn.close()

    def addText(self):
        db = self.openDB()
        cur = db[0]
        conn = db[1]
        cur.execute('select review_en from repeater where star<=3  ')
        text = cur.fetchall()
        print(len(text))
        print(text[1][0])
        t = '##'

        for i in range(len(text)):
            print(i)
            t= t+text[i][0]
        print(t)
        cur.execute('insert into text values (%s)', (t))
        conn.commit()
        cur.close()
        conn.close()







if __name__ == '__main__':
    app = wx.App()
    frame = ButtonFrame()
    frame.Show()
    app.MainLoop()




os.system("pause")
